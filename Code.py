############## MT20133 MT20123 OOPD PROJECT ##################

from openpyxl import *
from tkinter import *
from tkinter import messagebox
import datetime
import webbrowser

path='C:/Users/dell 1/Desktop/IIITD SEM1/OOPD/MT20133_MT20123_Project1_FS/oopd_db.xlsx'
#path='C:/Users/Admin/Desktop/Sem1/OOPD/project/oopd_db.xlsx'

wb = load_workbook(path)

################### ADVERTISMENT ########################

class Advertisement:
    '''
    Advertisement Class
    Responsibility --> Take advertisement info from administrator 
    '''
    def __init__(self,master,adInfo : str):
        self.master=master
        self._adInfo = adInfo
        self._adUI = AdvertisementUI(master,adInfo)

class AdvertisementUI:
    '''
    AdvertisementUI Class
    Responsibility --> Display UI for Adverstiment
    '''
    def __init__(self,master,adInfo):
        self.master = master
        master.title("ADVERTISEMENT")
        master.geometry("800x800")
        
        adFrame = Frame(master)
        adFrame.pack()

        self.heading = Label(adFrame, text = "CALL FOR ADMISSIONS!",font =("arial", 18))
        self.heading.pack()
        
        self.info=adInfo
        self.infoBox = Text(adFrame,wrap="none")
        self.infoBox.config(font =("arial", 14))
        self.infoBox.pack()
        self.infoBox.insert(END,self.info)
        
        self.weblink = Label(adFrame, text="Visit Institute Website", fg="blue", cursor="hand2")
        self.weblink.bind("<Button>", lambda e: webbrowser.open_new("https://iiitd.ac.in/"))
        self.weblink.pack(side=LEFT)
        
        self.closewindow = Button(width=10, bg="yellow",text="NEXT",command=master.destroy)
        self.closewindow.pack()


#################### APPLY FOR ADMISSION #########################

class ApplyForAdmissionPage:
    '''
    Apply for admission Class
    Responsibility --> Displaying the webpage for details of admission application 
    '''
    def __init__(self,master,applyInfo : str):
        self.master=master
        self._applyInfo = applyInfo
        self._applyUI = ApplyForAdmissionPageUI(master,applyInfo)

class ApplyForAdmissionPageUI:
    '''
    ApplyForAdmissionPageUI Class
    Responsibility --> Display UI for Apply for admission Page
    '''
    def __init__(self,master,applyInfo):
        self.master = master
        master.title("APPLY NOW")
        master.geometry("800x800")
        applyFrame = Frame(master)
        applyFrame.pack()

        self.heading = Label(applyFrame, text = "APPLY NOW!",font =("arial", 18))
        self.heading.pack()
        self.info=applyInfo
        self.infoBox = Text(applyFrame,wrap="none")
        self.infoBox.config(font =("arial", 14))
        self.infoBox.pack()
        self.infoBox.insert(END,self.info)
        
        self.weblink = Label(applyFrame, text="Visit Institute Website", fg="blue", cursor="hand2")
        self.weblink.bind("<Button>", lambda e: webbrowser.open_new("https://iiitd.ac.in/"))
        self.weblink.pack()
        
        self.closewindow = Button(width=10, bg="yellow",text="NEXT",command=master.destroy)
        self.closewindow.pack()


################  APPLICATION FORM #############

class ApplicationForm:
    '''
    ApplicationForm class
    Responsibility --> Display Application Form 
    '''
    def __init__(self,master):
        
        self.master = master
        self.applyUI = ApplicationFormUI(master)

class ApplicationFormDB:
    '''
    ApplicationFormDB class
    Responsibility --> Maintain the database for application form entries
    '''

    def __init__(self,mysheet):
        self.mysheet = mysheet

    def prepareSheet(self):
        self.mysheet.cell(row=1, column=1).value = "FullName"
        self.mysheet.cell(row=1, column=2).value = "DOB"
        self.mysheet.cell(row=1, column=3).value = "Contact"
        self.mysheet.cell(row=1, column=4).value = "Class10"
        self.mysheet.cell(row=1, column=5).value = "Class12"

    def SubmitForm(self,l1,l2,l3,l4,l5):
        current_row = self.mysheet.max_row 
        current_column = self.mysheet.max_column
        self.mysheet.cell(row=current_row + 1, column=1).value = l1.get()
        self.mysheet.cell(row=current_row + 1, column=2).value = l2.get() 
        self.mysheet.cell(row=current_row + 1, column=3).value = l3.get() 
        self.mysheet.cell(row=current_row + 1, column=4).value = l4.get()
        self.mysheet.cell(row=current_row + 1, column=5).value = l5.get() 
        wb.save(path)
        l1.focus_set()
        l1.delete(0, END) 
        l2.delete(0, END) 
        l3.delete(0, END)
        l4.delete(0, END) 
        l5.delete(0, END)
        messagebox.showinfo('Success','Thank you!.\nYour form has been submitted successfully.')

    def getDBaccess(self):
        return self.mysheet

        
class ApplicationFormUI:
    '''
    ApplicationFormUI Class
    Responsibility --> Display UI for Application Form
    '''
    def __init__(self,master):

        self.master = master
        master.geometry("1350x700+0+0")
        master.title("Application Form")
        master.configure(background="purple")
        self.heading=Label(master, text = "APPLICATION FORM",font =("times new roman",30,"bold"),bd=8,bg="navy",fg="white")
        self.heading.pack(side=TOP,fill=X)
    
        sideInstructionsFrame=Frame(master,bd=4,bg="cyan")
        sideInstructionsFrame.place(x=20,y=100,width=450,height=600)
    
        self.Instructions=Label(sideInstructionsFrame,text="Instructions for Candidates",font =("times new roman",20,"bold"),bg="cyan",fg="black")
        self.Instructions.grid(row=0,columnspan=2,pady=20)
        self.label_Ins1=Label(sideInstructionsFrame,text="1. All the entries in the form are mandatory",font =("times new roman",14),bg="cyan",fg="black")
        self.label_Ins1.grid(row=1,column=0,padx=20,pady=10,sticky="w")
        self.label_Ins2=Label(sideInstructionsFrame,text="2. Once you submit the form, you accept that all\n the information entered by you is True.",font =("times new roman",14),bg="cyan",fg="black")
        self.label_Ins2.grid(row=2,column=0,padx=20,pady=10,sticky="w")
        self.label_Ins3=Label(sideInstructionsFrame,text="2. For detailed instructions you can visit our website",font =("times new roman",14),bg="cyan",fg="black")
        self.label_Ins3.grid(row=3,column=0,padx=20,pady=10,sticky="w")
        self.link1 = Label(sideInstructionsFrame, text="Click here!", fg="blue", cursor="hand2",bg="cyan")
        self.link1.bind("<Button>", lambda e: webbrowser.open_new("https://iiitd.ac.in/"))
        self.link1.grid(row=4,column=0,padx=20,pady=0)
    
        FormFrame=Frame(master,bd=4,bg="cyan")
        FormFrame.place(x=500,y=100,width=820,height=600)
        self.label_f1=Label(FormFrame,text="Full Name",font =("times new roman",20),bg="cyan",fg="black")
        self.label_f1.grid(row=1,column=0,padx=20,pady=10,sticky="w")
        self.label_Entry1=Entry(FormFrame,font=("times new roman",15),bd=5)
        self.label_Entry1.grid(row=1,column=1,padx=20,pady=10,sticky="w")
        self.label_f2=Label(FormFrame,text="Date of Birth(DD/MM/YYYY)",font =("times new roman",20),bg="cyan",fg="black")
        self.label_f2.grid(row=2,column=0,padx=20,pady=10,sticky="w")
        self.label_Entry2=Entry(FormFrame,font=("times new roman",15),bd=5)
        self.label_Entry2.grid(row=2,column=1,padx=20,pady=10,sticky="w")
        self.label_f3=Label(FormFrame,text="Contact No.",font =("times new roman",20),bg="cyan",fg="black")
        self.label_f3.grid(row=3,column=0,padx=20,pady=10,sticky="w")
        self.label_Entry3=Entry(FormFrame,font=("times new roman",15),bd=5)
        self.label_Entry3.grid(row=3,column=1,padx=20,pady=10,sticky="w")
        self.label_f4=Label(FormFrame,text="Gender",font =("times new roman",20),bg="cyan",fg="black")
        self.label_f4.grid(row=4,column=0,padx=20,pady=10,sticky="w")
        var=IntVar()
        self.R1 = Radiobutton(FormFrame, text="Male", variable=var, value=1,command=var.get(),bg="cyan",font=("times new roman",15))
        self.R1.grid(row=4,column=1,padx=20,pady=10,sticky="w"  )
        self.R2 = Radiobutton(FormFrame, text="Female", variable=var, value=2, command=var.get(),bg="cyan",font=("times new roman",15))
        self.R2.grid(row=4,column=2,padx=10,pady=10,sticky="w"  )
        self.label_f5=Label(FormFrame,text="Class 10 CGPA(out of 10.0)",font =("times new roman",20),bg="cyan",fg="black")
        self.label_f5.grid(row=5,column=0,padx=20,pady=10,sticky="w")
        self.label_Entry5=Entry(FormFrame,font=("times new roman",15),bd=5)
        self.label_Entry5.grid(row=5,column=1,padx=20,pady=10,sticky="w")
        self.label_f6=Label(FormFrame,text="Class 12 Marks in %",font =("times new roman",20),bg="cyan",fg="black")
        self.label_f6.grid(row=6,column=0,padx=20,pady=10,sticky="w")
        self.label_Entry6=Entry(FormFrame,font=("times new roman",15),bd=5)
        self.label_Entry6.grid(row=6,column=1,padx=20,pady=10,sticky="w")

        self.label_f7=Button(FormFrame,text="Submit",font =("times new roman",20,"bold"),bg="light yellow",fg="black",cursor="hand2",command= lambda: self.validateInputAndSubmitForm())
        self.label_f7.grid(row=7,column=1,padx=20,pady=10,sticky="w")
        
    def validateInputAndSubmitForm(self):
        appDB = ApplicationFormDB(wb.active)
        
        self.flag = 1
        #Validation checks
        if (self.label_Entry1.get()=="" or self.label_Entry2.get()=="" or  self.label_Entry3.get()=="" or self.label_Entry5.get()=="" or self.label_Entry6.get()=="" ):
            self.flag=0
            messagebox.showinfo('Empty Entries', 'It is mandatory to fill each entry in the Form!\nPlease fill the form again.')

        if (len(self.label_Entry3.get())!=10):
            self.flag=0
            messagebox.showinfo('Invalid Data','The Contact number entered by you is invalid!\nPlease fill the form again.')
             
        try:
            datetime.datetime.strptime(self.label_Entry2.get(), "%d/%m/%Y")
        except:
            self.flag=0
            messagebox.showinfo('Invalid Data','The Date of Birth entered by you is invalid!\nPlease fill the form again.')

        if(self.flag==1):
            appDB.SubmitForm(self.label_Entry1,self.label_Entry2,self.label_Entry3,self.label_Entry5,self.label_Entry6)
        
        
################# CONDUCT ADMISSION TEST ##############

class eligibility:
    '''
    check eligibility for taking taking test Class
    Responsibility --> Take eligibility info from administrator and call UI
    '''
    def __init__(self,master):
        self.master=master
        self._adUI = eligibilityUI(master)
        
class eligibilityUI:
    def __init__(self,master):
        self.master=master
        appDB = ApplicationFormDB(wb.active)
        self.mysheet = appDB.getDBaccess()
        self.eligible_row=None
        master.title("SUBMIT DETAILS ")
        self.heading = Label(text= "Enter Details for the Test",font=("times new roman", 20))
        self.heading.grid(row=0,column=0,padx=20,pady=10)
        self.l1 =Label(text='Name:',font=("times new roman",16))
        self.l1.grid(row=1)
        self.l2=Label(text='Date of Birth:',font=("times new roman",16))
        self.l2.grid(row=2)
        self.name = Entry(font=("times new roman",16),bd=5)
        self.name.grid(row=1,column=1,padx=20,pady=10)
        self.dob = Entry(font=("times new roman",16),bd=5)
        self.dob.grid(row=2, column=1,padx=20,pady=18)
        self.submit=Button(text="Submit",font =("times new roman",15,"bold"),bg="light yellow",fg="black", \
                      command= lambda: self.check_eligibility(self.name,self.dob))
        self.submit.grid(row=3,column=1,padx=10,pady=10)

    def check_eligibility(self,name,dob):
        flag=1
        index_eligible_row = 1
        for row in self.mysheet.rows:
            if row[0].value == name.get() and row[1].value == dob.get():
                flag = 0
                self.eligible_row = index_eligible_row
                messagebox.showinfo('You are eligible for the test', 'Test is starting. All the Best!')
            index_eligible_row += 1
        if flag==1:
            messagebox.showinfo('Incorrect entry', 'Enter correct name and Date of birth\nPlease fill the form again.')

class take_test_eval:
    '''
    Take and evaluate Class
    Responsibility --> Take test info from administrator and call UI
    '''
    def __init__(self,master,questions,correct_answers,eligible_row,marks_per_ques):
        self.master=master
        self._adUI = take_test_eval_UI(master,questions,correct_answers,eligible_row,marks_per_ques)

class take_test_eval_UI:
    def __init__(self,master,questions,correct_answers,eligible_row,marks_per_ques):
        self.master = master
        appDB = ApplicationFormDB(wb.active)
        self.mysheet = appDB.getDBaccess()
        self.questions = questions
        options = ["True", "False"]
        master.title("TEST PAPER")
        self.heading = Label(text= 'Question Paper:'+str(marks_per_ques)+'marks/question',font =("times new roman", 18))
        self.heading.grid(row=0,column=0,padx=20,pady=10)
        
        self.var = [ IntVar() for i in range(len(questions))]
        self.ques = [None]*len(questions)
        self.radio1=[None]*len(questions)
        self.radio2=[None]*len(questions)
        for i in range(len(questions)):
            self.ques[i]=Label(text=questions[i],font =("times new roman",18))
            self.ques[i].grid(row=1+i,column=0,padx=20,pady=10)
            self.radio1[i]=Radiobutton(text= options[0], variable=self.var[i], value=1,command=self.var[i].get(),\
                                       font=("times new roman",15))
            self.radio1[i].grid(row=1+i,column=1,padx=20,pady=10)
            self.radio2[i]=Radiobutton(text=options[1], variable=self.var[i], value=2, command=self.var[i].get(),\
                                       font=("times new roman",15))
            self.radio2[i].grid(row=1+i,column=2,padx=10,pady=10)

        self.submit=Button(text="Submit",font =("times new roman",15,"bold"),bg="light yellow",fg="black", \
                      command= lambda: self.evaluate(self.var,correct_answers,marks_per_ques,eligible_row))
        self.submit.grid(row=5,column=1,padx=10,pady=10)
        
    def evaluate(self,var,correct_answers,marks_per_ques,eligible_row):
        responses =[i.get() for i in var]
        count = 0
        for i in range(len(correct_answers)):
            if responses[i]==correct_answers[i]:
                count = count + 1
        self.mysheet.cell(row=eligible_row,column=6).value = count*marks_per_ques
        wb.save(path)
        count = count*marks_per_ques
        messagebox.showinfo('Success','Thank you!.\n Your score is: '+ str(count)+'/'+str(marks_per_ques*len(self.questions)))


############## NOTICES ##########

class Notices:
    
    def __init__(self,master,title,btn,noticeInfo):

        self.master = master
        self.title = title
        self.btn = btn
        master.title(self.title)
        master.geometry("700x700")
        self.info=noticeInfo

        noticeFrame = Frame(master)
        noticeFrame.pack()
        self.heading = Label(noticeFrame, text = "NOTICE",font =("arial", 18))
        self.heading.pack()
        
        self.infoBox = Text(noticeFrame,wrap="none")
        self.infoBox.config(font =("arial", 14))
        self.infoBox.pack()
        self.infoBox.insert(END,self.info)
        
        self.weblink = Label(noticeFrame, text="Visit Institute Website", fg="blue", cursor="hand2")
        self.weblink.bind("<Button>", lambda e: webbrowser.open_new("https://iiitd.ac.in/"))
        self.weblink.pack(side=LEFT)
        
        self.closewindow = Button(width=10, bg="yellow",text=self.btn,command=master.destroy)
        self.closewindow.pack()

      
class InterviewNotice(Notices):
    '''
    Interview Notice Class
    Responsibility --> Post notice for interview 
    '''
    def __init__(self, master,noticeInfo : str):
        self.title = "NOTICE FOR INTERVIEW"
        self.btn = "NEXT"
        Notices.__init__(self,master,self.title,self.btn,noticeInfo)


class FeePaymentNotice(Notices):
    '''
    Fee Payment Notice Class
    Responsibility --> Post notice for Fee Payment
    '''
    def __init__(self, master,noticeInfo : str):
        
        self.title = "NOTICE FOR FEE PAYMENT FOR SEM-1"
        self.btn = "NEXT"
        Notices.__init__(self,master,self.title,self.btn,noticeInfo)

class TestNotice(Notices):
    '''
    Test Notice Class
    Responsibility --> Post notice for Test
    '''
    def __init__(self,master,noticeInfo : str):
        
        self.title = "NOTICE FOR ADMISSION TEST"
        self.btn = "NEXT"
        Notices.__init__(self,master,self.title,self.btn,noticeInfo)

class AdmissionLetterNotice(Notices):
    '''
    Admission Letter Notice Class
    Responsibility --> Print Final Admission Letter
    '''
    def __init__(self,master,noticeInfo : str):
        
        self.title = "ADMISSION LETTER"
        self.btn = "CLOSE"
        Notices.__init__(self,master,self.title,self.btn,noticeInfo)

############## INTERVIEW ####################

class CandidatesForInterview:
    def __init__(self,master,list_eligible_candidates):
        self.master = master
        appDB = ApplicationFormDB(wb.active)
        self.mysheet = appDB.getDBaccess()
        self.list = list_eligible_candidates

    def display_list_eligible_candidates(self):
        for i in range(2,self.mysheet.max_row + 1):
            x = self.mysheet.cell(row=i, column=6)
            if x.value is not None and x.value > 14:
                self.list.append((self.mysheet.cell(row=i, column=1).value))

        self.GUI = CandidatesForInterviewUI(self.master,self.list)

    def getCandidateList(self):
        return self.list

class CandidatesForInterviewUI:
    def __init__(self,master,clist):
        self.master = master
        self.clist = clist
        master.geometry("1000x700+0+0")
        master.title("List of candidates eligible for interview")
        master.configure(background='purple')
        
        self.label=Label(master,text="Congratulations! Following Candidates are eligible for interview",font =("times new roman",14,"bold"),bg="yellow",fg="black")
        self.label.grid(row=0,column=3,padx=20,pady=2)
        
        for i in range(0,len(self.clist)):
            self.label2=Label(master,text=self.clist[i],font =("times new roman",14,"bold"),bg="purple",fg="white")
            self.label2.grid(row=i+1,column=1,padx=20)
        


###############  MERIT LIST & WAIT LIST ##############

class Merit_Wait_List:
    def __init__(self,master,mlist,wlist,clist):
        self.master = master
        self.mlist = mlist
        self.wlist = wlist
        self.clist = clist

    def prepare_And_Display_Merit_List_With_Wait_List(self):
        #no. of seats=10
        for x in self.clist[:10]:
           if x is not None:
               self.mlist.append(x)

        for x in self.clist[11:16]:
           if x is not None:
               self.wlist.append(x)     
        #assume remaining do not clear interview

        self.GUI = Merit_Wait_List_UI(self.master,self.mlist,self.wlist)
        

class Merit_Wait_List_UI:
    def __init__(self,master,mlist,wlist):
        self.master = master
        self.mlist = mlist
        self.wlist = wlist
        master.geometry("1000x700+0+0")
        master.title("Merit/Wait List of Candidates:")
        master.configure(background='purple')

        self.label=Label(master,text="Congratulations ! Final Merit List:",font =("times new roman",18,"bold"),bg="yellow",fg="black")
        self.label.grid(row=0,column=0,padx=10,pady=10,sticky="w")
        for i in range(0,len(self.mlist)):
            self.label2=Label(master,text=self.mlist[i],font =("times new roman",14,"bold"),bg="purple",fg="white")
            self.label2.grid(row=i+1,column=1,sticky="w")

        self.label3=Label(master,text="Wait List:",font =("times new roman",18,"bold"),bg="yellow",fg="black")
        self.label3.grid(row=16,column=0,padx=10,pady=10,sticky="w")
        row_no=17
        for i in range(0,len(self.wlist)):
            self.label4=Label(master,text=self.wlist[i],font =("times new roman",14,"bold"),bg="purple",fg="white")
            self.label4.grid(row=row_no,column=1,sticky="w")
            row_no=row_no+1
    
            
def main():   

    root1 = Tk()
    adObject= Advertisement(root1,"\t\t\t20 June 2020\t\t\t\n\n\
                          Indraprastha Institute of Information Technlogy Delhi\n\n\
                          A State University established by Govt. of NCT Delhi\n\n\
                          Admission open for Academic Year 2020 \n\n\n\n\
                          IIITD stands best among the top engineering colleges in India\n\n\
                           Admissions open for 4 year B.Tech Program\n\n\n\
                           IMPORTANT DATES :\n\
                            1. Opening of Application Form: 6 July 2020\n\
                            2. Closing of Application Form: 15 July 2020\n\
                          \n\nFor further details contact admin@iiitd.ac.in \n or call on +919123456780 between 9am to 5pm \n\n")
    root1.mainloop()

    root2=Tk()
    applyObject = ApplyForAdmissionPage(root2,"\t\t\t6 July 2020\t\t\t\n\n  Application Process begins!\n\n Indraprastha Institute of Information Technlogy Delhi\n\n\
                           A State University established by Govt. of NCT Delhi\n\n\
                           Admission open for Academic Year 2020 \n\n\n\n\
                           IIITD stands best among the top engineering colleges in India\n\n\
                           Admissions open for 4 year B.Tech CSE Program\n\n\n\n\n\n\
                           For further details \n\n\
                           You can visit the website by clicking the following link:\n")
    root2.mainloop()

    root3=Tk()
    appObject = ApplicationForm(root3)
    root3.mainloop()

    root4=Tk()
    testObject = TestNotice(root4,"The online test will be conducted on 31 October,2020 at 14hours.\nThe duration of the test will be 30 minutes.\nCandidates are requested to be come live 10 minutes \nbefore the test.\nAny form of cheating is strictly prohibited.\n")
    root4.mainloop()

    root5 = Tk()
    adObject= eligibility(root5)
    root5.mainloop()
    eligible_row = adObject._adUI.eligible_row
    questions = ["1. The matrix [ [2, -1], [-2, 1] ] is invertible:", "2. The magnification produced by mirror is 0:",\
             "3. The medicine Paracetamol is an analgesic:", "4. Dennis Ritchie is the father of C language:"]
    correct_answers = [2,2,1,1]

    root6 = Tk()
    adObject= take_test_eval(root6,questions,correct_answers,eligible_row,5)
    root6.mainloop()

    root7=Tk()
    candidateInterviewObject = CandidatesForInterview(root7,[])
    candidateInterviewObject.display_list_eligible_candidates()
    candidatelist = candidateInterviewObject.getCandidateList()
    root7.mainloop()
    
    root8=Tk()
    intvObject = InterviewNotice(root8,"The shortlisted candidates are invited for the inteview.\nThe details are as follows:\nLocation: building II, college campus\nDate: 15 November,2020\nTime: 9 hours.\nBring your Aadhar card and resume.\n")
    root8.mainloop()

    root9=Tk()
    meritWaitObject = Merit_Wait_List(root9,[],[],candidatelist)
    meritWaitObject.prepare_And_Display_Merit_List_With_Wait_List()
    root9.mainloop()

    root10=Tk()
    feeObject = FeePaymentNotice(root10,"The fee details are as mentioned below.\nKindly pay the amount of Rs.80,000\n\n BANK NAME: ABC BANK OF INDIA\nIFSC CODE: 12XXXXXXX\nFor further details visit iiitd.ac.in\n")
    root10.mainloop()


    root11=Tk()
    finalletterObject = AdmissionLetterNotice(root11,"Welcome to IIITD!\n\n\
                                        You have been offered admission in \n\
                                        Indraprashtha Institute Of Information Technology for \n\
                                        B.Tech in Computer Science and Engineering\n\
                                        It is a delight to have you in our Institute and we promise\n\
                                        to deliver the best of knowledge\n\
                                        and skills for your bright future!")
    root11.mainloop()

if __name__ == "__main__":
    main()
